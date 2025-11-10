from rest_framework.parsers import MultiPartParser, FormParser
from django.core.exceptions import ValidationError
from .functions import calculate_file_hash, extract_boq
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import JsonResponse
from io import BytesIO
from .serializers import *
import json

from collections import defaultdict
from django.db.models import Q, Sum, Prefetch
from rest_framework.views import APIView
from rest_framework import generics
from rest_framework.response import Response
from django.db import transaction
from django.http import HttpResponse
import io
from .serializers import BOQItemCostUpdateSerializer
from estimation.models import BOQItem, Subsection, Section
from .serializers import BOQItemSerializer
from excel.serializers import AppConfigKV, AppConfigSerializer
from excel.permissions import RolePermission
from openpyxl import Workbook
from django.shortcuts import get_object_or_404
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict
from django.http import HttpResponse
from rest_framework.views import APIView
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import io



# Estimation Views
class EstimationCreateView(APIView):
    # authentication_classes = [TokenAuthentication]
    # permission_classes = [IsAuthenticated]

    def post(self, request, **kwargs):
        try:
            serializer = EstimationSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return JsonResponse({
                    "status": "success",
                    "message": "Estimation created successfully",
                    "data": serializer.data
                })
            return JsonResponse({
                "status": "failed",
                "message": serializer.errors,
                "data": None
            }, safe=False, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            print(f'Error: {e}')
            return JsonResponse({
                "status": "failed",
                "message": "Something went wrong...",
                "data": None
            }, safe=False, status=status.HTTP_400_BAD_REQUEST)


class EstimationUpdateView(APIView):
    # authentication_classes = [TokenAuthentication]
    # permission_classes = [IsAuthenticated]

    def put(self, request, pk, **kwargs):
        try:
            estimation = Estimation.objects.get(id=pk)
        except Estimation.DoesNotExist:
            return JsonResponse({
                "status": "failed",
                "message": "Estimation not found",
                "data": None
            }, safe=False, status=status.HTTP_404_NOT_FOUND)

        try:
            serializer = EstimationSerializer(estimation, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return JsonResponse({
                    "status": "success",
                    "message": "Estimation updated successfully",
                    "data": serializer.data
                })
            return JsonResponse({
                "status": "failed",
                "message": serializer.errors,
                "data": None
            }, safe=False, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            print(f'Error: {e}')
            return JsonResponse({
                "status": "failed",
                "message": "Something went wrong...",
                "data": None
            }, safe=False, status=status.HTTP_400_BAD_REQUEST)


class EstimationListView(APIView):
    # authentication_classes = [TokenAuthentication]
    # permission_classes = [IsAuthenticated]

    def get(self, request, **kwargs):
        try:
            project_id = request.query_params.get('project_id')
            estimations = Estimation.objects.filter(subphase__project_id=project_id)
            serializer = EstimationSerializer(estimations, many=True)
            return JsonResponse({
                "status": "success",
                "message": None,
                "data": serializer.data
            })
        except Exception as e:
            print(f'Error: {e}')
            return JsonResponse({
                "status": "failed",
                "message": "Something went wrong...",
                "data": None
            }, safe=False, status=status.HTTP_400_BAD_REQUEST)

class UploadBOQ(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('boq_file')
        estimation_id = request.data.get('estimation_id')

        if not file:
            return Response({'detail': 'BOQ file is required.'}, status=status.HTTP_400_BAD_REQUEST)

        if not estimation_id:
            return Response({'detail': 'Estimation ID is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            estimation = Estimation.objects.get(id=estimation_id)
        except Estimation.DoesNotExist:
            return Response({'detail': 'Estimation not found.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            file_hash = calculate_file_hash(file)

            # Check if a BOQ with same file hash already exists
            existing_boq = BOQ.objects.filter(estimation=estimation, file_hash=file_hash).first()
            if existing_boq:
                return Response({
                    'detail': 'Duplicate BOQ file found. Skipping extraction.',
                    'boq_id': existing_boq.id
                }, status=status.HTTP_200_OK)

            # Create and extract new BOQ
            boq = BOQ.objects.create(
                name=f"BOQ for {estimation.subphase.name}",
                estimation=estimation,
                file_path=file,
                file_hash=file_hash
            )

            extraction_success = extract_boq(file, boq)
            if extraction_success:
                return Response({
                    'detail': 'BOQ data extracted and saved successfully.',
                    'boq_id': boq.id
                }, status=status.HTTP_201_CREATED)
            else:
                return Response({
                    'detail': 'File upload succeeded but data extraction failed.',
                    'boq_id': boq.id
                }, status=status.HTTP_400_BAD_REQUEST)

        except ValidationError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            print(f"Error in UploadBOQ: {e}")
            return Response({'detail': 'Unexpected error occurred.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class BOQListView(APIView):
    # authentication_classes = [TokenAuthentication]
    # permission_classes = [IsAuthenticated]

    def get(self, request, **kwargs):
        try:
            estimation_id = request.query_params.get('estimation_id')
            boqs = BOQ.objects.filter(estimation__subphase__project_id=estimation_id)
            data = BOQSerializer(boqs, many=True).data

            return JsonResponse({
                "status": "success",
                "message": None,
                "data": data
            })
        except Exception as e:
            print(f'Error: {e}')
            return JsonResponse({
                "status": "failed",
                "message": "Something went wrong...",
                "data": None
            }, safe=False, status=status.HTTP_400_BAD_REQUEST)


class BOQDetailView(APIView):
    # authentication_classes = [TokenAuthentication]
    # permission_classes = [IsAuthenticated]

    def get(self, request, pk, **kwargs):
        try:
            # Fetch the specific BOQ by ID
            boq = BOQ.objects.get(id=pk)

            # Serialize the data for BOQ, Sections, Subsections, and BOQItems
            boq_serializer = BOQDetailSerializer(boq)
            # sections = Section.objects.filter(boq=boq)
            # subsections = Subsection.objects.filter(section__in=sections)
            # boq_items = BOQItem.objects.filter(subsection__in=subsections)

            # sections_serializer = SectionSerializer(sections, many=True)
            # subsections_serializer = SubsectionSerializer(subsections, many=True)
            # boq_items_serializer = BOQItemSerializer(boq_items, many=True)

            return JsonResponse({
                "status": "success",
                "message": None,
                "data": boq_serializer.data,
            })
        except BOQ.DoesNotExist:
            return JsonResponse({
                "status": "failed",
                "message": "BOQ not found",
                "data": None
            }, safe=False, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f'Error: {e}')
            return JsonResponse({
                "status": "failed",
                "message": "Something went wrong...",
                "data": None
            }, safe=False, status=status.HTTP_400_BAD_REQUEST)


class BOQUpdateView(APIView):
    def put(self, request, pk, **kwargs):
        try:
            # Fetch the specific BOQ by ID
            boq = BOQ.objects.get(id=pk)

            # Save the current values for revision logging
            original_boq = BOQSerializer(boq)
            original_sections = Section.objects.filter(boq=boq)
            original_subsections = Subsection.objects.filter(section__in=original_sections)
            original_boq_items = BOQItem.objects.filter(subsection__in=original_subsections)

            # Create a list of changes that will be logged in the BOQRevision table
            revisions = []
            revision_fields = {}

            # Update the BOQ itself (name or other fields)
            if 'name' in request.data:
                new_name = request.data['name']
                if boq.name != new_name:
                    revision_fields["name"] = {
                        "old_value": boq.name,
                        "new_value": new_name
                    }
                    boq.name = new_name  # Update the BOQ name
                    boq.save()  # Save the updated BOQ

            # Iterate through the sections and check for changes
            if 'sections' in request.data:
                section_data = request.data['sections']
                for data in section_data:
                    section = Section.objects.get(id=data['id'])
                    if section.name != data['name']:
                        revision_fields["section"] = {
                            "old_value": section.name,
                            "new_value": data['name']
                        }
                        section.name = data['name']
                        section.save()

            # Iterate through the subsections and check for changes
            if 'subsections' in request.data:
                subsection_data = request.data['subsections']
                for data in subsection_data:
                    subsection = Subsection.objects.get(id=data['id'])
                    if subsection.name != data['name']:
                        revision_fields["subsection"] = {
                            "old_value": subsection.name,
                            "new_value": data['name']
                        }
                        subsection.name = data['name']
                        subsection.save()

            # Iterate through the BOQItems and check for changes
            if 'boq_items' in request.data:
                boq_item_data = request.data['boq_items']
                for boq_item in original_boq_items:
                    for data in boq_item_data:
                        if data.get('id') == boq_item.id:
                            # Prepare the old values and new values for BOQItem fields
                            old_boq_item = {
                                "description": boq_item.description,
                                "unit": boq_item.unit,
                                "quantity": str(boq_item.quantity),  # Ensure it's a string
                                "rate": str(boq_item.rate),
                                "amount": str(boq_item.amount),
                                "subsection_id": boq_item.subsection.id  # Save only the ID of Subsection
                            }
                            new_boq_item = {
                                "description": data.get('description'),
                                "unit": data.get('unit'),
                                "quantity": str(data.get('quantity', 0)),
                                "rate": str(data.get('rate', 0)),
                                "amount": str(data.get('amount', 0)),
                                "subsection_id": data.get('subsection', boq_item.subsection.id)  # Save only the ID of Subsection
                            }

                            # If any fields have changed, log the change as a single revision entry
                            if old_boq_item != new_boq_item:
                                # Log the old and new values for BOQItem
                                revision_fields["boq_item"] = {
                                    "old_value": old_boq_item,
                                    "new_value": new_boq_item
                                }

                            # Update the BOQItem
                            boq_item.description = data.get('description')
                            boq_item.unit = data.get('unit')
                            boq_item.quantity = data.get('quantity')
                            boq_item.rate = data.get('rate')
                            boq_item.amount = data.get('amount')

                            # Update the related foreign key for the subsection if it changed
                            if data.get('subsection') != boq_item.subsection.id:
                                boq_item.subsection = Subsection.objects.get(id=data.get('subsection'))

                            boq_item.save()

            # Now log the changes in a single BOQRevision entry
            if revision_fields:
                revision_data = {
                    "boq": boq,
                    "old_value": json.dumps(revision_fields),
                    "new_value": json.dumps(revision_fields),
                    "updated_at": timezone.now()
                }
                BOQRevision.objects.create(**revision_data)

            return JsonResponse({
                "status": "success",
                "message": "BOQItem(s), Section(s), Subsection(s) updated successfully",
                "data": "Updated BOQItems, Sections, and Subsections successfully"
            })
        except BOQ.DoesNotExist:
            return JsonResponse({
                "status": "failed",
                "message": "BOQ not found",
                "data": None
            }, safe=False, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f'Error: {e}')
            return JsonResponse({
                "status": "failed",
                "message": "Something went wrong...",
                "data": None
            }, safe=False, status=status.HTTP_400_BAD_REQUEST)


# ---------- Grouped + Subtotals + GrandTotal (+addons) ----------
class BOQGroupedView(APIView):
    """
    Groups BOQ items by Section and calculates subtotals and grand totals,
    including tax and overheads.
    """
    def get(self, request):
        # fetch non-deleted items with optimized joins
        items = (
            BOQItem.objects
            .filter(is_deleted=False)
            .select_related("subsection__section")
        )

        cfg, _ = AppConfigKV.objects.get_or_create(pk=1)

        # skip unwanted descriptions (case-insensitive)
        unwanted = ["collection", "total brought forward from page no.", "carried to final summary"]
        q_unwanted = Q()
        for u in unwanted:
            q_unwanted |= Q(description__iexact=u)
        items = items.exclude(q_unwanted)

        # group by section
        grouped = defaultdict(list)
        for it in items:
            section_name = (
                it.subsection.section.name
                if it.subsection and it.subsection.section
                else "Uncategorized"
            )
            grouped[section_name].append(it)

        sections = []
        grand = 0.0
        for sec_name, rows in grouped.items():
            # derive section_factor from first item or default 0
            section_factor = rows[0].factor if rows else 0.0

            serialized = BOQItemSerializer(rows, many=True).data

            subtotal = sum(float(r.get("amount") or 0) for r in serialized)
            subtotal_with_factor = subtotal * (1 + (section_factor or 0.0) / 100)
            grand += subtotal_with_factor

            sections.append({
                "name": sec_name,
                "subtotal": subtotal_with_factor,
                "section_factor": section_factor,
                "items": serialized
            })

        tax = grand * (cfg.tax_rate or 0)
        overhead = grand * (cfg.overhead_rate or 0)
        grand_total = grand + tax + overhead

        return Response({
            "sections": sections,
            "baseTotal": grand,
            "tax": tax,
            "overhead": overhead,
            "grandTotal": grand_total,
            "config": AppConfigSerializer(cfg).data
        })


# ---------- Update one item (inline) ----------
class BOQItemUpdateView(generics.RetrieveUpdateAPIView):
    queryset = BOQItem.objects.all()
    serializer_class = BOQItemSerializer
    permission_classes = [RolePermission]


# ---------- Create/Delete/Restore/Hard-delete ----------
class BOQItemCreateView(generics.CreateAPIView):
    serializer_class = BOQItemSerializer
    permission_classes = [RolePermission]


class BOQItemSoftDeleteView(APIView):
    def delete(self, request, pk):
        try:
            item = BOQItem.objects.get(pk=pk)
            item.soft_delete(reason=request.query_params.get("reason", ""))
            return Response(status=204)
        except BOQItem.DoesNotExist:
            return Response(status=404)


class BOQItemRestoreView(APIView):
    def post(self, request, pk):
        try:
            item = BOQItem.objects.get(pk=pk)
            item.restore()
            return Response(status=200)
        except BOQItem.DoesNotExist:
            return Response(status=404)


class BOQItemHardDeleteView(generics.DestroyAPIView):
    queryset = BOQItem.objects.all()
    permission_classes = [RolePermission]


# ---------- Reorder inside a subsection ----------
class ReorderView(APIView):
    """
    Payload: { "subsection_id": 12, "ids": [5,3,9,...] }
    """
    def post(self, request):
        subsection_id = request.data.get("subsection_id")
        ids = request.data.get("ids", [])
        if not subsection_id or not isinstance(ids, list):
            return Response({"detail": "Invalid payload"}, status=400)

        with transaction.atomic():
            for idx, _id in enumerate(ids, start=1):
                BOQItem.objects.filter(pk=_id, subsection_id=subsection_id).update(sort_order=idx)
        return Response({"message": "Reordered"}, status=200)


class ExportExcelView(APIView):
    """
    Exports BOQ hierarchy (Section → Subsection → Items) to Excel
    with grouped headers for Materials, Plant, Labour, Subcontract,
    and tail summary fields.
    Handles multi-material rows cleanly without repeating item data.
    """

    def get(self, request, boq_id):
        # 1️⃣ Get BOQ
        try:
            boq = BOQ.objects.get(pk=boq_id)
        except BOQ.DoesNotExist:
            return Response({"detail": "BOQ not found"}, status=404)

        # 2️⃣ Prefetch related data
        item_qs = BOQItem.objects.filter(is_deleted=False).order_by("sort_order", "id").prefetch_related("materials")
        subsection_qs = Subsection.objects.prefetch_related(
            Prefetch("items", queryset=item_qs, to_attr="prefetched_items")
        )
        sections = (
            Section.objects
            .filter(boq=boq)
            .order_by("id")
            .prefetch_related(
                Prefetch("subsections", queryset=subsection_qs, to_attr="prefetched_subsections")
            )
        )

        cfg, _ = AppConfigKV.objects.get_or_create(pk=1)

        # 3️⃣ Workbook setup
        wb = Workbook()
        ws = wb.active
        ws.title = "BOQ Export"

        # Branding Header
        ws.merge_cells("A1:W1")
        ws["A1"] = cfg.brand_title or f"Bill of Quantities - {boq.name}"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:W2")
        subtitle = cfg.brand_subtitle or (boq.estimation.subphase.project.name if boq.estimation else "")
        ws["A2"] = subtitle
        ws["A2"].alignment = Alignment(horizontal="center")

        # 4️⃣ Header setup (row3 = group, row4 = subheaders)
        ws.append([])
        ws.append([])
        header_row = 3
        subheader_row = 4

        def set_cell(row, col, value, bold=False, align="center"):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            if bold:
                cell.font = Font(bold=True)
            return cell

        # Fixed columns
        fixed_cols = ["Item No", "Description", "Unit", "Quantity", "Rate", "Amount"]
        for idx, title in enumerate(fixed_cols, start=1):
            set_cell(header_row, idx, title, bold=True)
            ws.merge_cells(start_row=header_row, start_column=idx,
                           end_row=subheader_row, end_column=idx)

        col = 7

        # MATERIAL (5 cols now: Name, Rate, Wastage, U/Rate, Amount)
        ws.merge_cells(start_row=header_row, start_column=col, end_row=header_row, end_column=col + 4)
        set_cell(header_row, col, "MATERIAL", bold=True)
        subheaders = ["Name", "Rate", "Wastage", "U/Rate", "Amount"]
        for i, sub in enumerate(subheaders):
            set_cell(subheader_row, col + i, sub)
        col += 5

        # PLANT (2 cols)
        ws.merge_cells(start_row=header_row, start_column=col, end_row=header_row, end_column=col + 1)
        set_cell(header_row, col, "PLANT", bold=True)
        set_cell(subheader_row, col, "Rate")
        set_cell(subheader_row, col + 1, "Amount")
        col += 2

        # LABOUR (2 cols)
        ws.merge_cells(start_row=header_row, start_column=col, end_row=header_row, end_column=col + 1)
        set_cell(header_row, col, "LABOUR", bold=True)
        set_cell(subheader_row, col, "Hrs.")
        set_cell(subheader_row, col + 1, "Amount")
        col += 2

        # SUBCONTRACT (2 cols)
        ws.merge_cells(start_row=header_row, start_column=col, end_row=header_row, end_column=col + 1)
        set_cell(header_row, col, "SUBCONTRACT", bold=True)
        set_cell(subheader_row, col, "Rate")
        set_cell(subheader_row, col + 1, "Amount")
        col += 2

        # Tail fields
        tail_cols = ["Dry Cost", "UNIT RATE", "Prelimin.", "BOQ Amount", "Amount"]
        for title in tail_cols:
            set_cell(header_row, col, title, bold=True)
            ws.merge_cells(start_row=header_row, start_column=col,
                           end_row=subheader_row, end_column=col)
            col += 1

        thin = Side(border_style="thin", color="999999")
        row_idx = 5
        base_total = 0.0

        # 5️⃣ Data Rows
        for section in sections:
            ws.append([None])
            ws.cell(row=row_idx, column=2, value=f"SECTION - {section.name}")
            ws.cell(row=row_idx, column=2).font = Font(bold=True, size=11)
            row_idx += 1

            for subsection in getattr(section, "prefetched_subsections", []):
                ws.append([None])
                ws.cell(row=row_idx, column=2, value=subsection.name)
                ws.cell(row=row_idx, column=2).font = Font(italic=True)
                row_idx += 1

                for item in getattr(subsection, "prefetched_items", []):
                    materials = list(item.materials.all())
                    plant_rate = float(item.plant_rate or 0)
                    plant_amount = float(item.plant_amount or 0)
                    labour_hrs = float(item.labor_hours or 0)
                    labour_amount = float(item.labor_amount or 0)
                    subcontract_rate = float(item.subcontract_rate or 0)
                    subcontract_amount = float(item.subcontract_amount or 0)
                    dry_cost = float(item.dry_cost or 0)
                    unit_rate = float(item.unit_rate or 0)
                    prelimin = float(item.prelimin or 0)
                    boq_amount = float(item.boq_amount or 0)
                    tail_amount = float(item.amount or 0)
                    main_amount = float(item.amount or 0)
                    base_total += main_amount

                    # --- main BOQItem row ---
                    ws.append([
                        item.id,
                        f"{item.description or ''}",
                        item.unit or "",
                        float(item.quantity or 0),
                        float(item.rate or 0),
                        main_amount,
                        "", "", "", "", "",  # 5 material cols blank
                        plant_rate,
                        plant_amount,
                        labour_hrs,
                        labour_amount,
                        subcontract_rate,
                        subcontract_amount,
                        dry_cost,
                        unit_rate,
                        prelimin,
                        boq_amount,
                        tail_amount,
                    ])
                    for c in range(1, 24):
                        ws.cell(row=row_idx, column=c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
                    row_idx += 1

                    # --- materials (sub-rows) ---
                    for m in materials:
                        ws.append([
                            "",  # Item No blank
                            "",  # Description blank (removed ↳ name here)
                            "", "", "", "",  # Unit–Amount blanks
                            m.name or "",  # MATERIAL Name
                            m.rate or "",
                            m.wastage or "",
                            m.u_rate or "",
                            float(m.amount or 0),
                            "", "", "", "", "", "", "", "", "", "", "", "", ""
                        ])
                        for c in range(1, 24):
                            ws.cell(row=row_idx, column=c).border = Border(
                                top=thin, bottom=thin, left=thin, right=thin
                            )
                        row_idx += 1

                ws.append([])
                row_idx += 1

            ws.append(["", "Carried to Collection"])
            for c in range(1, 24):
                ws.cell(row=row_idx, column=c).font = Font(bold=True)
                ws.cell(row=row_idx, column=c).fill = PatternFill("solid", fgColor="FFF4D7")
            row_idx += 2

        # 6️⃣ Totals
        tax = base_total * (cfg.tax_rate or 0)
        overhead = base_total * (cfg.overhead_rate or 0)
        grand = base_total + tax + overhead

        ws.append([])
        row_idx += 1
        totals = [
            ("Base Total:", base_total),
            (f"Tax ({(cfg.tax_rate or 0)*100:.1f}%)", tax),
            (f"Overhead ({(cfg.overhead_rate or 0)*100:.1f}%)", overhead),
            ("Grand Total:", grand),
        ]
        start_col = 19
        for label, value in totals:
            ws.cell(row=row_idx, column=start_col, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=start_col + 1, value=value).font = Font(bold=True)
            row_idx += 1

        for c in range(start_col, start_col + 2):
            ws.cell(row=row_idx - 1, column=c).fill = PatternFill("solid", fgColor="E2F0D9")

        # 7️⃣ Autosize
        for i in range(1, 25):
            column_letter = get_column_letter(i)
            max_len = 12
            for cell in ws[column_letter]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_len + 2, 60)

        # 8️⃣ Return Excel
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f"attachment; filename=BOQ_{boq_id}_Hierarchical.xlsx"
        return resp


class ExportPDFView(APIView):
    """
    Export BOQ (hierarchical) to PDF — styled A3 landscape layout
    matching the Excel exporter with all columns visible.
    """

    def get(self, request, boq_id):
        # ----- Fetch BOQ -----
        try:
            boq = BOQ.objects.get(pk=boq_id)
        except BOQ.DoesNotExist:
            return Response({"detail": "BOQ not found"}, status=404)

        # Prefetch optimized data
        item_qs = BOQItem.objects.filter(is_deleted=False).order_by("sort_order", "id").prefetch_related("materials")
        subsection_qs = Subsection.objects.prefetch_related(
            Prefetch("items", queryset=item_qs, to_attr="prefetched_items")
        )
        sections = (
            Section.objects.filter(boq=boq)
            .order_by("id")
            .prefetch_related(
                Prefetch("subsections", queryset=subsection_qs, to_attr="prefetched_subsections")
            )
        )

        cfg, _ = AppConfigKV.objects.get_or_create(pk=1)

        # ----- PDF Setup -----
        buf = BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A3),
            leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20,
        )

        styles = getSampleStyleSheet()
        wrap = ParagraphStyle("wrap", parent=styles["Normal"], fontSize=6, leading=7, wordWrap="CJK")
        wrap_bold = ParagraphStyle("wrap_bold", parent=wrap, fontName="Helvetica-Bold")

        story = []

        # ----- Branding Header -----
        if cfg.logo_url:
            try:
                story.append(RLImage(cfg.logo_url, width=70, height=70))
            except Exception:
                pass
        story.append(Paragraph(f"<b>{cfg.brand_title or f'Bill of Quantities - {boq.name}'}</b>", styles["Title"]))
        subtitle = cfg.brand_subtitle or (boq.estimation.subphase.project.name if boq.estimation else "")
        if subtitle:
            story.append(Paragraph(subtitle, styles["Normal"]))
        story.append(Spacer(1, 10))

        # ----- Table Header -----
        headers = [
            "Item No", "Description", "Unit", "Qty", "Rate", "Amount",
            "Mat Name", "Mat Rate", "Wastage", "U/Rate", "Mat Amt",
            "Plant Rate", "Plant Amt",
            "Lab Hrs", "Lab Amt",
            "Sub Rate", "Sub Amt",
            "Dry Cost", "UNIT RATE", "Prelimin.", "BOQ Amt", "Amount"
        ]
        data = [headers]
        total_amount = 0.0

        # ----- Hierarchical Rows -----
        for section in sections:
            data.append(["", Paragraph(f"<b>{section.name}</b>", wrap_bold)] + [""] * (len(headers) - 2))
            for subsection in getattr(section, "prefetched_subsections", []):
                data.append(["", Paragraph(f"<b>{subsection.name}</b>", wrap_bold)] + [""] * (len(headers) - 2))

                for item in getattr(subsection, "prefetched_items", []):
                    total_amount += float(item.amount or 0)

                    # Main item row
                    data.append([
                        item.id,
                        Paragraph(item.description or "", wrap),
                        item.unit or "",
                        f"{float(item.quantity or 0):g}",
                        f"{float(item.rate or 0):g}",
                        f"{float(item.amount or 0):,.2f}",
                        "", "", "", "", "",
                        f"{float(item.plant_rate or 0):g}",
                        f"{float(item.plant_amount or 0):,.2f}",
                        f"{float(item.labor_hours or 0):g}",
                        f"{float(item.labor_amount or 0):,.2f}",
                        f"{float(item.subcontract_rate or 0):g}",
                        f"{float(item.subcontract_amount or 0):,.2f}",
                        f"{float(item.dry_cost or 0):,.2f}",
                        f"{float(item.unit_rate or 0):,.2f}",
                        f"{float(item.prelimin or 0):,.2f}",
                        f"{float(item.boq_amount or 0):,.2f}",
                        f"{float(item.amount or 0):,.2f}",
                    ])

                    # Material subrows (gray background)
                    for m in item.materials.all():
                        row = [
                            "", "", "", "", "", "",
                            m.name or "",
                            f"{m.rate or 0:g}",
                            f"{m.wastage or 0:g}",
                            f"{m.u_rate or 0:g}",
                            f"{float(m.amount or 0):,.2f}",
                        ] + [""] * (len(headers) - 11)
                        data.append(row)

            data.append(["", Paragraph("<b>Carried to Collection</b>", wrap_bold)] + [""] * (len(headers) - 2))
            data.append([""] * len(headers))

        # ----- Table Styling -----
        col_widths = [
            30, 110, 35, 35, 40, 45,  # core
            60, 40, 40, 40, 50,       # materials
            40, 45,                   # plant
            40, 45,                   # labour
            40, 45,                   # subcontract
            45, 45, 45, 50, 50        # totals
        ]
        table = Table(data, repeatRows=1, colWidths=col_widths)

        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E1F2")),  # header background
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#000000")),
            ("FONTSIZE", (0, 0), (-1, -1), 6),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.whitesmoke, colors.white]),  # alternate subtle
        ]))

        # Highlight material subrows
        for r_idx, row in enumerate(data):
            if len(row) > 6 and row[6] and not row[1]:  # has material name, no description
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, r_idx), (-1, r_idx), colors.HexColor("#F7F7F7")),
                ]))

        story.append(table)
        story.append(Spacer(1, 12))

        # ----- Totals -----
        tax = total_amount * (cfg.tax_rate or 0)
        overhead = total_amount * (cfg.overhead_rate or 0)
        grand = total_amount + tax + overhead

        totals = [
            ["Base Total", f"{total_amount:,.2f}"],
            [f"Tax ({(cfg.tax_rate or 0)*100:.1f}%)", f"{tax:,.2f}"],
            [f"Overhead ({(cfg.overhead_rate or 0)*100:.1f}%)", f"{overhead:,.2f}"],
            ["Grand Total", f"{grand:,.2f}"],
        ]
        t = Table(totals, colWidths=[220, 120])
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E2F0D9")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
        ]))
        story.append(t)

        # ----- Build PDF -----
        doc.build(story)
        buf.seek(0)
        response = HttpResponse(buf, content_type="application/pdf")
        response["Content-Disposition"] = f"attachment; filename=BOQ_{boq_id}_Detailed.pdf"
        return response



# ---------- MATERIAL ----------
class MaterialCreateView(generics.CreateAPIView):
    serializer_class = MaterialSerializer
    queryset = Material.objects.all()


class MaterialUpdateView(generics.UpdateAPIView):
    serializer_class = MaterialSerializer
    queryset = Material.objects.all()
    lookup_field = "pk"


class MaterialByBOQItemView(APIView):
    def get(self, request, boq_item_id):
        materials = Material.objects.filter(boq_item_id=boq_item_id)
        serializer = MaterialSerializer(materials, many=True)
        return Response(serializer.data)


class BOQItemCostUpdateView(APIView):
    """
    Unified API to update BOQItem costs:
    Supports labour, plant, and subcontract cost additions/updates in a single endpoint.
    """

    def post(self, request, boq_item_id):
        """
        Accepts JSON payload like:
        {
            "labor_hours": 5,
            "labor_rate": 120,
            "plant_rate": 200,
            "subcontract_rate": 350
        }
        Only provided fields will be updated.
        """
        boq_item = get_object_or_404(BOQItem, id=boq_item_id)

        serializer = BOQItemCostUpdateSerializer(
            boq_item, data=request.data, partial=True
        )
        if serializer.is_valid():
            serializer.save()
            return Response({
                "message": "BOQItem costs updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)