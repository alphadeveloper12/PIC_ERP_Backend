import io
import pandas as pd
from collections import defaultdict
from django.db import transaction
from django.http import HttpResponse
from rest_framework import status, generics
from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Q
from .models import BillItem, AppConfigKV, Project, Revision, Material
from .serializers import *
from .permissions import RolePermission
from django.db.models import Q
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet
from .functions import _norm_cols, choose_sheet_fast, _alias_df, to_float_series, BULK_SIZE, REPLACE_MATERIALS


class ProjectView(APIView):
    # Fetch and create projects
    def get(self, request):
        projects = Project.objects.all()
        serializer = ProjectSerializer(projects, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = ProjectSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class RevisionView(APIView):
    # Manage project revisions
    def post(self, request):
        project = Project.objects.get(id=request.data.get('project_id'))
        revision_version = project.revisions.count() + 1  # Auto-increment version number
        revision = Revision.objects.create(
            project=project,
            version=revision_version,
            file=request.FILES['file']
        )
        return Response(RevisionSerializer(revision).data, status=status.HTTP_201_CREATED)


class UploadBillView(APIView):
    def post(self, request):
        project_id = request.data.get('project_id')
        if not project_id:
            return Response({"error": "project_id is required"}, status=status.HTTP_400_BAD_REQUEST)


        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            return Response({"error": f"Project with id {project_id} does not exist."},
                            status=status.HTTP_404_NOT_FOUND)

        if 'file' not in request.FILES:
            return Response({"error": "file is required"}, status=status.HTTP_400_BAD_REQUEST)

        fobj = request.FILES['file']

        # 1) Pick sheet + header
        try:
            sheet_name, header_row = choose_sheet_fast(fobj)
        except Exception as e:
            return Response({"error": f"Failed to inspect Excel: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        # 2) Load chosen sheet fully
        try:
            fobj.seek(0)
        except Exception:
            pass
        try:
            df = pd.read_excel(fobj, sheet_name=sheet_name, header=header_row, dtype=str)
        except Exception as e:
            return Response({"error": f"Failed to read chosen sheet: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        # 3) Normalize/alias headers
        df.columns = _norm_cols(df.columns)
        df = _alias_df(df)

        # 4) Filter rows + detect sections
        desc = df.get("description")
        if desc is None:
            return Response({"error": "No 'description' column found in the selected sheet."},
                            status=status.HTTP_400_BAD_REQUEST)

        desc_clean = desc.astype(str).str.strip()
        valid_mask = (desc_clean.ne("")) & (desc_clean.str.lower().ne("nan"))
        skip_mask = desc_clean.str.lower().str.contains(r"(total brought forward|collection)", regex=True)

        df = df[valid_mask & ~skip_mask].copy()
        if df.empty:
            return Response({"error": "No usable rows after filtering."}, status=status.HTTP_400_BAD_REQUEST)

        df["__is_section__"] = desc_clean.loc[df.index].str.upper().str.startswith("SECTION")
        df.loc[df["__is_section__"], "__section_val__"] = desc_clean.loc[df["__is_section__"].index]
        df["__section_val__"] = df["__section_val__"].ffill()

        items_df = df[~df["__is_section__"]].copy()
        if items_df.empty:
            return Response({"error": "Only section rows found; no items to import."}, status=status.HTTP_400_BAD_REQUEST)

        # 5) Coerce numerics vectorized; ensure columns exist
        num_cols = [
            "quantity","rate","material_rate","material_wastage","plant_rate",
            "labour_hours","subcontract_rate","others_1_qty","others_1_rate",
            "others_2_qty","others_2_rate","dry_cost","unit_rate","prelimin",
            "boq_amount","actual_qty","actual_amount","factor"
        ]
        for c in num_cols:
            if c in items_df.columns:
                items_df[c] = to_float_series(items_df[c])
            else:
                items_df[c] = 0.0

        # Ensure text fields present
        for c in ["unit", "item_no", "material"]:
            if c not in items_df.columns:
                items_df[c] = ""

        # Normalize key fields + sort_order within section
        items_df["section"] = items_df["__section_val__"].fillna("Uncategorized")
        items_df["description"] = desc_clean.loc[items_df.index]
        items_df["item_no"] = items_df["item_no"].fillna("").astype(str).replace({"nan": ""})
        items_df["unit"] = items_df["unit"].fillna("").astype(str)
        items_df["sort_order"] = items_df.groupby("section").cumcount() + 1

        # 6) Prepare flat records for DB layer
        common_fields = [
            "section","sort_order","item_no","description","unit","quantity","rate",
            "material_rate","material_wastage","plant_rate","labour_hours","subcontract_rate",
            "others_1_qty","others_1_rate","others_2_qty","others_2_rate",
            "dry_cost","unit_rate","prelimin","boq_amount","actual_qty","actual_amount","factor"
        ]
        materials_series = items_df["material"].fillna("").astype(str)

        records = items_df[common_fields].to_dict("records")

        # Build sets for bulk fetch
        sections = list(items_df["section"].unique())
        item_nos = list(items_df["item_no"].unique())
        descriptions = list(items_df["description"].unique())

        # 7) Bulk fetch existing items and split create/update (NO UPSERT)
        try:
            existing_qs = BillItem.objects.filter(
                project=project,
                section__in=sections,
                item_no__in=item_nos,
                description__in=descriptions
            ).only("id","section","item_no","description")
        except Exception as e:
            return Response({"error": f"Failed to fetch existing items: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        # Key normalization used throughout
        def k(section, item_no, description):
            return ((section or "Uncategorized"), (item_no or ""), (description or ""))

        existing_map = {k(bi.section, bi.item_no, bi.description): bi for bi in existing_qs}

        to_create, to_update = [], []
        update_fields = [
            "sort_order","unit","quantity","rate",
            "material_rate","material_wastage","plant_rate","labour_hours","subcontract_rate",
            "others_1_qty","others_1_rate","others_2_qty","others_2_rate",
            "dry_cost","unit_rate","prelimin","boq_amount","actual_qty","actual_amount","factor"
        ]

        # Pre-bucket materials by key
        material_bucket = {}
        for idx, rec in enumerate(records):
            key = k(rec["section"], rec["item_no"], rec["description"])
            material_text = materials_series.iloc[idx]
            names = [m.strip() for m in str(material_text).split(";") if m.strip()]
            material_bucket[key] = names

            exists = existing_map.get(key)
            if exists:
                # copy fields to existing instance
                for f in update_fields + ["section","item_no","description"]:
                    setattr(exists, f, rec.get(f))
                to_update.append(exists)
            else:
                to_create.append(
                    BillItem(project=project, **rec)
                )

        # 8) Persist everything in one transaction
        with transaction.atomic():
            if to_create:
                BillItem.objects.bulk_create(to_create, batch_size=BULK_SIZE)
                for bi in to_create:
                    existing_map[k(bi.section, bi.item_no, bi.description)] = bi

            if to_update:
                BillItem.objects.bulk_update(to_update, update_fields, batch_size=BULK_SIZE)

            # materials
            affected_ids = [bi.id for bi in existing_map.values()]
            if REPLACE_MATERIALS and affected_ids:
                Material.objects.filter(bill_item_id__in=affected_ids).delete()

            mat_objs = []
            for key, names in material_bucket.items():
                bi = existing_map.get(key)
                if not bi or not names:
                    continue
                for nm in names:
                    mat_objs.append(Material(bill_item_id=bi.id, material_name=nm))
                    if len(mat_objs) >= BULK_SIZE:
                        Material.objects.bulk_create(mat_objs, batch_size=BULK_SIZE)
                        mat_objs = []
            if mat_objs:
                Material.objects.bulk_create(mat_objs, batch_size=BULK_SIZE)

            # Optional heavy recompute of amounts via model hook:
            # for bi in existing_map.values():
            #     bi.save(force_calculation=True)

        return Response(
            {"message": "Uploaded & saved with estimations.", "sheet_used": sheet_name},
            status=status.HTTP_201_CREATED
        )


class ConfigView(APIView):
    def get(self, request):
        cfg, _ = AppConfigKV.objects.get_or_create(pk=1)
        return Response(AppConfigSerializer(cfg).data)

    def put(self, request):
        cfg, _ = AppConfigKV.objects.get_or_create(pk=1)
        ser = AppConfigSerializer(cfg, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data)


# ---------- Flat list (non-deleted) ----------
class BillItemListView(generics.ListAPIView):
    serializer_class = BillItemSerializer

    def get_queryset(self):
        project_id = self.request.query_params.get('project_id')
        return BillItem.objects.filter(project_id=project_id, is_deleted=False)


# ---------- Grouped + Subtotals + GrandTotal (+addons) ----------
class BillGroupedView(APIView):
    def get(self, request):
        # Fetch non-deleted BillItems
        items = BillItem.objects.filter(is_deleted=False)
        cfg, _ = AppConfigKV.objects.get_or_create(pk=1)

        # Filter out unwanted rows like "COLLECTION" and "Total Brought Forward"
        unwanted_descriptions = [
            "collection",
            "total brought forward from page no.",
            "carried to final summary"
        ]
        
        # Build Q object to filter based on multiple descriptions
        q_filter = Q()
        for desc in unwanted_descriptions:
            q_filter |= Q(description__iexact=desc)

        # Exclude items matching any of the unwanted descriptions
        items = items.exclude(q_filter)

        # Group items by section
        grouped = defaultdict(list)
        for it in items:
            grouped[it.section or "Uncategorized"].append(it)

        sections = []
        grand = 0.0
        for sec, rows in grouped.items():
            # Fetch section factor (from BillItem or Section model)
            section_factor = rows[0].section_factor if rows else 0.0  # Assuming section_factor is stored in BillItem

            # Serialize the BillItems in the section
            data = BillItemSerializer(rows, many=True).data

            # Calculate subtotal
            subtotal = sum((r["amount"] or 0) for r in data)
            
            # Apply section factor to subtotal
            subtotal_with_factor = subtotal * (1 + section_factor)
            
            grand += subtotal_with_factor

            # Append section details
            sections.append({
                "name": sec,
                "subtotal": subtotal_with_factor,
                "section_factor": section_factor,  # Include section factor in response
                "items": data
            })

        # Calculate totals
        tax = grand * (cfg.tax_rate or 0)
        overhead = grand * (cfg.overhead_rate or 0)
        grand_total = grand + tax + overhead

        return Response({
            "sections": sections,
            "baseTotal": grand,
            "tax": tax,
            "overhead": overhead,
            "grandTotal": grand_total,
            "config": AppConfigSerializer(cfg).data
        })

# ---------- Update one item (inline) ----------
class BillItemUpdateView(generics.RetrieveUpdateAPIView):
    queryset = BillItem.objects.all()
    serializer_class = BillItemSerializer
    permission_classes = [RolePermission]


# ---------- Create/Delete/Restore/Hard-delete ----------
class BillItemCreateView(generics.CreateAPIView):
    serializer_class = BillItemSerializer
    permission_classes = [RolePermission]

class BillItemSoftDeleteView(APIView):
    def delete(self, request, pk):
        try:
            item = BillItem.objects.get(pk=pk)
            item.soft_delete(reason=request.query_params.get("reason", ""))
            return Response(status=204)
        except BillItem.DoesNotExist:
            return Response(status=404)

class BillItemRestoreView(APIView):
    def post(self, request, pk):
        try:
            item = BillItem.objects.get(pk=pk)
            item.restore()
            return Response(status=200)
        except BillItem.DoesNotExist:
            return Response(status=404)

class BillItemHardDeleteView(generics.DestroyAPIView):
    queryset = BillItem.objects.all()
    permission_classes = [RolePermission]


# ---------- Reorder inside a section ----------
class ReorderView(APIView):
    """
    Payload: { "section": "SECTION B - SITE WORK", "ids": [5,3,9, ...] }
    """
    def post(self, request):
        section = request.data.get("section")
        ids = request.data.get("ids", [])
        if not section or not isinstance(ids, list):
            return Response({"detail": "Invalid payload"}, status=400)

        with transaction.atomic():
            for idx, _id in enumerate(ids, start=1):
                BillItem.objects.filter(pk=_id, section=section).update(sort_order=idx)
        return Response({"message": "Reordered"}, status=200)


# ---------- Export Excel with subtotals & grand total ----------
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ExportExcelView(APIView):
    def get(self, request):
        items = BillItem.objects.filter(is_deleted=False)
        cfg, _ = AppConfigKV.objects.get_or_create(pk=1)

        grouped = defaultdict(list)
        for it in items:
            grouped[it.section or "Uncategorized"].append(it)

        wb = Workbook()
        ws = wb.active
        ws.title = "BOQ"

        # Brand header
        ws.merge_cells("A1:G1")
        ws["A1"] = cfg.brand_title or "Bill of Quantities"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:G2")
        ws["A2"] = cfg.brand_subtitle or ""
        ws["A2"].alignment = Alignment(horizontal="center")

        # header row
        headers = ["Section", "Item No", "Description", "Unit", "Quantity", "Rate", "Amount"]
        ws.append([])
        ws.append(headers)
        for c in range(1, len(headers)+1):
            ws.cell(row=3, column=c).font = Font(bold=True)

        thin = Side(border_style="thin", color="999999")
        row_idx = 4
        base_total = 0.0

        for sec, rows in grouped.items():
            ws.append([sec])
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            row_idx += 1

            subtotal = 0.0
            for r in rows:
                ws.append([
                    "", r.item_no, r.description, r.unit,
                    r.quantity or 0, r.rate or 0, r.amount or 0
                ])
                for c in range(1, 8):
                    ws.cell(row=row_idx, column=c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
                subtotal += (r.amount or 0)
                row_idx += 1

            ws.append(["", "", "", "", "", "Subtotal:", subtotal])
            for c in range(1, 8):
                ws.cell(row=row_idx, column=c).font = Font(bold=True)
                ws.cell(row=row_idx, column=c).fill = PatternFill("solid", fgColor="FFF4D7")
            row_idx += 1
            base_total += subtotal

            ws.append([])
            row_idx += 1

        tax = base_total * (cfg.tax_rate or 0)
        overhead = base_total * (cfg.overhead_rate or 0)
        grand = base_total + tax + overhead

        ws.append(["", "", "", "", "", "Base Total:", base_total]); row_idx += 1
        ws.append(["", "", "", "", "", f"Tax ({cfg.tax_rate*100:.1f}%)", tax]); row_idx += 1
        ws.append(["", "", "", "", "", f"Overhead ({cfg.overhead_rate*100:.1f}%)", overhead]); row_idx += 1
        ws.append(["", "", "", "", "", "Grand Total:", grand]); row_idx += 1
        for c in range(1, 8):
            ws.cell(row=row_idx, column=c).font = Font(bold=True)
            ws.cell(row=row_idx, column=c).fill = PatternFill("solid", fgColor="E2F0D9")

        # autosize
        for col in ws.columns:
            max_len = 12
            letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[letter].width = min(max_len + 2, 60)

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = 'attachment; filename=BOQ_with_Subtotals.xlsx'
        return resp


# ---------- Export PDF (branded) ----------
class ExportPDFView(APIView):
    def get(self, request):
        items = BillItem.objects.filter(is_deleted=False)
        cfg, _ = AppConfigKV.objects.get_or_create(pk=1)
        grouped = defaultdict(list)
        for it in items:
            grouped[it.section or "Uncategorized"].append(it)

        base_total = 0.0
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
        story = []
        styles = getSampleStyleSheet()

        header_flow = []
        if cfg.logo_url:
            try:
                header_flow.append(RLImage(cfg.logo_url, width=80, height=80))
            except:
                pass
        header_flow.append(Paragraph(f"<b>{cfg.brand_title or 'Bill of Quantities'}</b>", styles['Title']))
        if cfg.brand_subtitle:
            header_flow.append(Paragraph(cfg.brand_subtitle, styles['Normal']))
        story += header_flow + [Spacer(1, 12)]

        for sec, rows in grouped.items():
            story.append(Paragraph(sec, styles['Heading2']))
            data = [["Item No", "Description", "Unit", "Qty", "Rate", "Amount"]]
            subtotal = 0.0
            for r in rows:
                data.append([
                    r.item_no or "",
                    r.description or "",
                    r.unit or "",
                    f"{r.quantity or 0:g}",
                    f"{r.rate or 0:g}",
                    f"{r.amount or 0:g}",
                ])
                subtotal += (r.amount or 0)
            base_total += subtotal
            data.append(["", "", "", "", "Subtotal", f"{subtotal:g}"])

            tbl = Table(data, repeatRows=1, colWidths=[60, 380, 60, 60, 60, 80])
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#e8e8e8")),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("GRID", (0,0), (-1,-1), 0.3, colors.grey),
                ("ALIGN", (3,1), (-1,-1), "RIGHT"),
                ("BACKGROUND", (-2,-1), (-1,-1), colors.HexColor("#fff4d7")),
                ("FONTNAME", (-2,-1), (-1,-1), "Helvetica-Bold"),
            ]))
            story += [tbl, Spacer(1, 12)]

        tax = base_total * (cfg.tax_rate or 0)
        overhead = base_total * (cfg.overhead_rate or 0)
        grand = base_total + tax + overhead

        totals = [
            ["Base Total", f"{base_total:g}"],
            [f"Tax ({(cfg.tax_rate or 0)*100:.1f}%)", f"{tax:g}"],
            [f"Overhead ({(cfg.overhead_rate or 0)*100:.1f}%)", f"{overhead:g}"],
            ["Grand Total", f"{grand:g}"],
        ]
        t = Table(totals, colWidths=[200, 120])
        t.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.3, colors.grey),
            ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#E2F0D9")),
            ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
            ("ALIGN", (1,0), (-1,-1), "RIGHT"),
        ]))
        story += [Spacer(1, 8), t]
        doc.build(story)

        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type="application/pdf")
        resp['Content-Disposition'] = 'attachment; filename=BOQ_Report.pdf'
        return resp
